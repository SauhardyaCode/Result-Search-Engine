import openpyxl

path = "C:\\Users\\USER\\letuscode\\Exploration\\result 12 data\\results.xlsx"

obj = openpyxl.load_workbook(path)
sheet = obj.active

for i in range(2, 153):
    marks = []
    for j in range(3, 18):
        cell = sheet.cell(i,j).value
        if str(cell).isnumeric():
            marks.append(cell)

    if len(marks)>5:
        marks.sort()
        marks.pop(0)

    sheet.cell(i, 18).value = sum(marks)
    sheet.cell(i, 19).value = sum(marks)/5
    obj.save("C:\\Users\\USER\\letuscode\\Exploration\\result 12 data\\faadu.xlsx")