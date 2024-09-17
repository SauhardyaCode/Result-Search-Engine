import openpyxl, os
from tkinter import *
from PIL import Image, ImageTk

path = "C:\\Users\\USER\\letuscode\\Exploration\\result 12 data"

obj = openpyxl.load_workbook(path+"/faadu.xlsx")
sheet = obj.active

root = Tk()
root.state('zoomed')
root.title("Result Analytics")

heading = Label(root, text="Result Search Engine", fg="#16423C", font="Helvetica 25 bold")

frame = Frame(root)
pregunta = Label(frame, text="Search by name or roll...", fg="#603F26", font="Comicsans 15")

search = StringVar()
respuesta = Entry(frame, width="30", font="Verdana 20", textvariable=search)
respuesta.focus_set()

glass_size = 50
glass_img = Image.open(path+"/glass.png")
glass_img = glass_img.resize((glass_size,glass_size))
glass_img = ImageTk.PhotoImage(glass_img)
glass = Button(frame, image=glass_img, width=glass_size, height=glass_size, bd=0, activebackground="#7FA1C3")

heading.pack(ipady=20)
frame.pack(pady=30)
pregunta.grid(row=0,column=0, pady=10)
respuesta.grid(row=1, column=0, pady=20, ipady=8)
glass.grid(row=1,column=1, ipadx=10)

suggestions_frame = Frame(frame)
suggestions_frame.grid(row=2, column=0)

valid = False
prediction = None

options = []
def searching(ev=None):
    global valid, options, index, prediction
    index = -1
    value = search.get().strip()
    for wgt in suggestions_frame.winfo_children():
        wgt.destroy()
    if value:
        options = []
        suggestions = []
        if value.isnumeric():
            for i in range(2, 153):
                roll = str(sheet.cell(i, 1).value)
                if value in roll:
                    suggestions.append(roll)
                    if len(suggestions)>=8:
                        break
        else:
            for i in range(2, 153):
                name = sheet.cell(i, 2).value
                if value.upper() in name:
                    suggestions.append(name)
                    if len(suggestions)>=8:
                        break

        for i in range(len(suggestions)):
            btn = Button(suggestions_frame, text=suggestions[i], bd=0, font="Arial 15", background="#ffffff", width=20, command=lambda i=i:click(suggestions[i].lower()))
            btn.grid(row=2+i,column=0, ipadx=145)
            btn.bind("<Down>", down)
            btn.bind("<Up>", up)
            btn.bind("<Return>", lambda ev, i=i:click(suggestions[i].lower()))
            btn.bind("<Key>", lambda ev: respuesta.focus_set())
            options.append(btn)
        valid = True
    
        if not suggestions:
            label = Label(suggestions_frame, text="No Results!", fg="red", font="Arial 12")
            label.grid(row=2,column=0)
            valid = False
            prediction = None
        if search.get().upper() not in suggestions:
            valid = False
            try:
                prediction = suggestions[0]
            except IndexError:
                pass

def click(data, ev=None):
    global valid
    search.set(data)
    valid = True
    respuesta.focus_set()
    respuesta.icursor(END)
    searching()

def result(ev=None):
    global prediction, valid
    data = search.get().strip().upper()
    if data:
        root.destroy()
        win = Tk()
        win.state('zoomed')

        if not valid and prediction:
            data = prediction
            valid = True
        if valid:
            subjects = [sheet.cell(1,j).value[:-6] for j in range(3,18)]
            codes = [sheet.cell(1,j).value[-4:-1] for j in range(3,18)]

            if data.isnumeric():
                for i in range(2, 153):
                    if sheet.cell(i, 1).value == int(data):
                        resultado = [sheet.cell(i, j).value for j in range(1,20)]
            else:
                for i in range(2, 153):
                    if sheet.cell(i, 2).value == data:
                        resultado = [sheet.cell(i, j).value for j in range(1,20)]

            win.title(f"Result @ {resultado[0]}")

            Label(win, text="CENTRAL BOARD OF SECONDARY EDUCATION", fg="#2E073F", font="Courier 25 bold").pack(ipady=20)
            Label(win, text="MARKS STATEMENT CUM CERTIFICATE", fg="#2E073F", font="Courier 18 bold").pack(ipady=10)
            Label(win, text="SENIOR SCHOOL CERTIFICATE EXAMINATION, 2024", fg="#2E073F", font="Courier 18 bold").pack(ipady=10)

            identity_frame = Frame(win)
            identity_frame.pack(padx=50, pady=40, anchor=W)

            Label(identity_frame, text="This is to certify that", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=0, ipadx=10, ipady=5, sticky=W)
            Label(identity_frame, text=resultado[1], font="Helvetica 12 bold").grid(row=0, column=1, ipadx=10, ipady=5, sticky=W)
            Label(identity_frame, text="Roll No.", fg="#603F26", font=('Times New Roman',15)).grid(row=1, column=0, ipadx=10, ipady=5, sticky=W)
            Label(identity_frame, text=resultado[0], font="Helvetica 12 bold").grid(row=1, column=1, ipadx=10, ipady=5, sticky=W)
            Label(identity_frame, text="School", fg="#603F26", font=('Times New Roman',15)).grid(row=2, column=0, ipadx=10, ipady=5, sticky=W)
            Label(identity_frame, text="19248 RAILWAY HIGHER SEC SCH N F RAILWAY  ALIPURDUAR WB", font="Helvetica 12 bold").grid(row=2, column=1, ipadx=10, ipady=5, sticky=W)
            Label(win, text="has achieved Scholastic Achievements as under :", fg="#603F26", font=('Times New Roman',15)).pack(ipadx=10, anchor=W, padx=50)

            results_frame = Frame(win, highlightbackground="#603F26", highlightthickness=2)
            results_frame.pack(padx=50, pady=10)

            popped_codes = []
            popped_subjects = []
            for i, m in enumerate(resultado[2:-2]):
                if m==None:
                    popped_codes.append(codes[i])
                    popped_subjects.append(subjects[i])

            for p in popped_codes:
                codes.remove(p)
            for p in popped_subjects:
                subjects.remove(p)

            Label(results_frame, text="SUB. CODE", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=0, ipadx=10, ipady=5)
            for i in range(len(codes)):
                Label(results_frame, text=codes[i], font="Helvetica 12 bold").grid(row=i+2, column=0, ipady=5)
            Label(results_frame, text="SUBJECT", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=1, ipadx=10, ipady=5)
            for i in range(len(codes)):
                Label(results_frame, text=subjects[i], font="Helvetica 12 bold").grid(row=i+2, column=1, ipady=5, sticky=W)

            marks = [x for x in resultado[2:-2] if x!=None]
            
            Label(results_frame, text="MARKS OBTAINED", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=2, ipadx=10, ipady=5)
            Label(results_frame, text="POSITIONAL GRADE", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=3, ipadx=10, ipady=5)
            for i,m in enumerate(marks):
                if m>90:grade='A1'
                elif m>80:grade='A2'
                elif m>70:grade='B1'
                elif m>60:grade='B2'
                elif m>50:grade='C1'
                elif m>40:grade='C2'
                elif m>32:grade='D1'
                elif m>20:grade='D2'
                else:grade='E'
                if m<100:m='0'+str(m)
                Label(results_frame, text=m, font="Helvetica 12 bold").grid(row=i+2, column=2, ipady=5)
                Label(results_frame, text=grade, font="Helvetica 12 bold").grid(row=i+2, column=3, ipady=5)
            
            conclusion_frame = Frame(win)
            conclusion_frame.pack(padx=50, anchor=E)

            Label(conclusion_frame, text="Total Marks", fg="#603F26", font=('Times New Roman',15)).grid(row=0, column=0, ipadx=10, ipady=5, sticky=W)
            Label(conclusion_frame, text=resultado[-2], font="Helvetica 12 bold").grid(row=0, column=1, ipadx=10, ipady=5, sticky=W)
            Label(conclusion_frame, text="Percentage", fg="#603F26", font=('Times New Roman',15)).grid(row=1, column=0, ipadx=10, sticky=W)
            Label(conclusion_frame, text=resultado[-1], font="Helvetica 12 bold").grid(row=1, column=1, ipadx=10, sticky=W)

        else:
            win.title("Error")
            Label(win, text="NO RECORDS FOUND!", fg="red", font="Courier 30 bold").pack(expand=1)

        back_size = 40
        back_img = Image.open(path+"/back.png")
        back_img = back_img.resize((back_size,back_size))
        back_img = ImageTk.PhotoImage(back_img)
        back = Button(win, image=back_img, bd=0, command=lambda:restart(win))
        back.place(x=0,y=0)

        win.mainloop()

def restart(win):
    os.startfile(path+"/analysis.pyw")
    win.destroy()

index = -1
def down(ev=None):
    global options, index
    index+=1
    selection()

def up(ev=None):
    global options, index
    index-=1
    selection()

def selection():
    global index, options
    if index>=len(options):
        index=0
        respuesta.focus_set()
    if index<0:
        index = len(options)-1
    options[index].focus_set()


glass.config(command=result)
respuesta.bind("<Return>", result)
respuesta.bind("<KeyRelease>", searching)
respuesta.bind("<Down>", down)
respuesta.bind("<Up>", up)

root.mainloop()